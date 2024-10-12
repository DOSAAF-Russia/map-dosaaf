import json
import random
import re
from datetime import date, datetime
from pprint import pprint

import aiohttp
import bs4
import geopy as gp
import httpx
import numpy as np
import orjson
import pandas as pd
import ua_generator
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from playwright.async_api import async_playwright
from playwright_stealth.stealth import stealth_async
from yarl import URL

from map_dosaaf.backend.database import prepare_db
from map_dosaaf.backend.database.repos import ECRepository, OrganisationRepository
from map_dosaaf.backend.utils import get_sqlalchemy_async_sessionmaker
from map_dosaaf.common.app_types import EC, Organisation


def organisations_iter():
    df = pd.read_csv("data/Сводная таблица по организациям ДОСААФ.csv")
    df = df.replace({np.nan: None})
    for i, row in df.iterrows():
        emails = row["Контактная почта организации"]
        if pd.isnull(emails):
            emails = []
        else:
            emails = emails.split(",")

        phones = row["Контактный(е) телефон(ы) организации"]
        if pd.isnull(phones):
            phones = []
        else:
            phones = phones.split(",")

        name = row["Название"]
        full_name = row["Полное название организации"]
        state = row["Состояние"]
        leader_ein = row["ИНН Руководителя"]
        leader_fio = row["Руководитель"]
        ein = row["ИНН Организации"]
        ogrn = row["ОГРН"]
        date_registration = datetime.strptime(row["Дата регистрации"], "%d.%m.%Y")
        contact_emails = emails
        contact_phones = phones
        geo_info = json.loads(row["Информация о расположении"])

        obj = Organisation(
            name=row["Название"],
            full_name=row["Полное название организации"],
            state=row["Состояние"],
            ein=str(row["ИНН Организации"]),
            ogrn=str(row["ОГРН"]),
            date_registration=datetime.strptime(row["Дата регистрации"], "%d.%m.%Y"),
            contact_emails=emails,
            contact_phones=phones,
            geo_info=json.loads(row["Информация о расположении"]),
        )

        if isinstance(leader_ein, (int, float)):
            obj.leader_ein = str(int(leader_ein))
        if isinstance(leader_fio, str):
            obj.leader_fio = leader_fio

        type = _transform_object_name(obj.full_name)

        yield obj


async def csv2db():
    await prepare_db()
    # df = pd.read_csv("data/Сводная таблица по организациям ДОСААФ.csv")
    # df = pd.read_json("list-org-gathered2-coords.json")
    df = pd.read_json("data/Юридические-Организации-list_org.json")
    df = df.replace({np.nan: None})
    maker = get_sqlalchemy_async_sessionmaker()
    async with maker() as session:
        repo = OrganisationRepository(session)
        # for obj in organisations_iter():
        for i, row in df.iterrows():
            ...
            obj = Organisation(
                full_name=row["full_name"],
                state=row["status"],
                ein=str(row["ein"]),
                contact_emails=row["emails"],
                contact_phones=row["phones"],
                websites=row["websites"],
                address=row["address"],
                coords=row["coords"],
                type_org=row["type_org"],
                federal_district=row["geolevels"]["2"],
                region=row["geolevels"]["3"],
                link_list_org=row["link"],
            )

            # coords = f"{obj.geo_info["lat"]}, {obj.geo_info["lon"]}"
            # print(coords)
            # pprint(obj.model_dump())

            await repo.add(obj)


async def parse_regional_departments():
    url = "https://www.dosaaf.ru/about/regions/"
    client = httpx.AsyncClient()

    resp = await client.get(url)
    html = resp.text

    soup = bs4.BeautifulSoup(html, "lxml")
    data = []

    for tag in soup.find_all("div", class_="ot-title-block_region"):
        tag: bs4.Tag

        fd = tag.find("h2").text
        regions = [
            t.text
            for t in tag.find_next_sibling("div")
            .find("div", class_="shortcode-content")
            .find_all("li")
        ]
        obj = {"name": fd, "regions": [{"name": n} for n in regions]}
        data.append(obj)

    return data


abbreviations = {
    "АШ": ["Автошкола", "Автомобильная школа"],
    "РТСК": "Республиканский спортивно-технический комплекс",
    "МО": ["Местная организация", "Местное отделение"],
    "ЗЕЦ": "Зональный единый центр",
    "ОТШ": "Объединённая техническая школа",
    "УЦ": "Учебный центр",
    "МЦ": "Мобильный центр",
    "СТК": [
        "Спортивно-технический комплекс",
        "СПОРТИВНО-ТЕХНИЧЕСКИЙ КЛУБ",
        "СПОРТИВНО - ТЕХНИЧЕСКИЙ КЛУБ",
        "СПОРТИВНО ТЕХНИЧЕСКИЙ КЛУБ"
    ],
    "СК": "Спортивный комплекс",
    "ТШ": "Техническая школа",
    "СШ": "Спортивная Школа",
    "КУСЦ": "Краевой учебно-спортивный центр",
    "СПАК": "Спортивно-парашютный авиационный клуб",
    # "АНО": "Автономная некоммерческая организация",
    "ССЦ": ["СТРЕЛКОВО-СПОРТИВНЫЙ ЦЕНТР", "СПОРТИВНО-СТРЕЛКОВЫЙ ЦЕНТР"],
    "ДПО": "ДОПОЛНИТЕЛЬНОГО ПРОФЕССИОНАЛЬНОГО ОБРАЗОВАНИЯ",
    "СТЦ": [
        "Спортивно-технический центр",
        "СПОРТИВНО-ТРЕНИРОВОЧНЫЙ ЦЕНТР",
        "СПОРТИВНО-ТРЕНИРОВОЧНЫЙ ЦЕНТР",
    ],
    "ЦППВ": "Центр профессиональной подготовки и патриотического воспитания",
    "ЦВПВ": "Центр военно-патриотического воспитания",
    "ВПЦ": "ВОЕННО-ПАТРИОТИЧЕСКИЙ ЦЕНТР",
    "МШ": "Морская школа",
    "ШСП": "Школа специальной подготовки",
    "ВПВ": "Военно-патриотического воспитания",
    "АК": "Аэроклуб",
    "ТЦ": "Технический центр",
    "ЦСС": "Центр служебного собаководства",
    "КОТШ": "Камчатская объединенная техническая школа",
    "ГО": "Городское отделение",
    "ОУ ДПО": "Образовательное учреждение дополнительного профессионального образования",
    "ССК": [
        "СПОРТИВНЫЙ СТРЕЛКОВЫЙ КЛУБ",
        "СТРЕЛКОВО-СПОРТИВНЫЙ КЛУБ",
        "СПОРТИВНО-СТРЕЛКОВЫЙ КЛУБ",
    ],
    "УАЦ": ["УЧЕБНЫЙ АВИАЦИОННЫЙ ЦЕНТР", "Учебно-авиационный центр"],
    "АУСЦ": "Авиационный учебно-спортивный центр",
    "АСК": ["АВИАЦИОННО-СПОРТИВНЫЙ КЛУБ", 'АВИАЦИОННЫЙ СПОРТИВНЫЙ КЛУБ'],
    "УСЦ": ["учебно-спортивный центр", "УЧЕБНО - СПОРТИВНЫЙ ЦЕНТР"],
    "УМЦ": "УЧЕБНО-МЕТОДИЧЕСКИЙ ЦЕНТР",
    "УСТК": "УЧЕБНЫЙ СПОРТИВНО-ТЕХНИЧЕСКИЙ КЛУБ",
    "ЗЦ": ["Зональный центр", "Зональный многофункциональный центр"],
    "РО": ["Региональное отделение", "Районное отделение"],
    "РЦ": [
        "Региональный центр",
        "Региональный многофункциональный центр",
        "Региональный единый центр",
    ],
    "ЕМЦ": "Единый многофункциональный центр",
    "ЕЦ": "Единый центр",
    "АТСК": "АВИАЦИОННО-ТЕХНИЧЕСКИЙ СПОРТИВНЫЙ КЛУБ",
    "СТТ": "СПОРТИВНО-СТРЕЛКОВЫЙ ТИР",
    
    "ФСУ": "ФИЗКУЛЬТУРНО-СПОРТИВНОЕ УЧРЕЖДЕНИЕ",
    "АРЗ": "АВИАЦИОННО-РЕМОНТНЫЙ ЗАВОД",
    "ПОУ": "ПРОФЕССИОНАЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ",
    # "": "",
    # "": "",
}


def _transform_object_name(incorrect_word: str):
    correct_word = " ".join([w.strip() for w in incorrect_word.split()])
    for key, value in abbreviations.items():
        if isinstance(value, str):
            # if any(correct_word.lower().count(w) for w in [f"{key.lower()} ", f" {key.lower()}"]):
            #     return correct_word.replace(key, value), value
            if correct_word.lower().count(value.lower()):
                return correct_word, value.capitalize()

        elif isinstance(value, list):
            for val in value:
                # if any(correct_word.lower().count(w) for w in [f"{key.lower()} ", f" {key.lower()}"]):
                #     return correct_word.replace(key, val), val
                if correct_word.lower().count(val.lower()):
                    return correct_word, val.capitalize()

        else:
            raise ValueError(value)

    return correct_word, None


async def get_types():
    types = []
    raw_types = []
    for org in organisations_iter():
        type_org = _transform_object_name(org.full_name)[1]

        types.append({"name": org.full_name, "type": type_org})
        raw_types.append(type_org)

    return types


class CheckoAPI:
    def __init__(self, proxies: list[str] = None) -> None:
        self._proxies = proxies
        self._client = aiohttp.ClientSession()
        self._base_url = "https://checko.ru"

    async def __aenter__(self):
        if self._client.closed:
            self._client = aiohttp.ClientSession()
        return self

    async def __aexit__(self, *args):
        await self._client.close()

    @property
    def _hdrs(self):
        headers = {
            "sec-ch-ua": '"Chromium";v="128", "Not;A=Brand";v="24", "Google Chrome";v="128"',
            "Referer": "https://checko.ru/",
            "sec-ch-ua-mobile": "?0",
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36",
            "sec-ch-ua-platform": '"Linux"',
        }
        return headers

    async def _query(self, q: str, proxy: str = None):
        url = f"{self._base_url}/search/quick_tips?query={q}"
        # return print(proxy)
        hdrs = {
            "accept": "*/*",
            "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7,my;q=0.6",
            "cache-control": "no-cache",
            # 'cookie': 'privacy_accept=true; _ym_uid=172624079269430310; _ym_isad=2; page_view_count=5; _checko_ru_session=UtJLPXDKvApbdQJ5NhPxcRbL7apH%2FuKpoc9qBl8IGGXGhNCML2irvaD9f13u388t0P0meaKbfXEMJ3X7lXqQ6kAzr4s6rWel3DDG05gNOGn0HBdf8elD0BQf2HmaFi9Aij4FrlWR%2FHW%2F9uh%2BHc13YNno%2BKAe3TN376xi0dpufLPRWGdhoJAk26n%2FZzrEm8fe0ieyCx%2B6mkTwj%2BEqtc5MzIQxxqQOd2H6IP14VKEgWipbBt7%2FzPKxY13vLtnOQtDNwo7uVXidOMlBVXPg%2BA5BDhkKFdw60VB8cGY%3D--VJKgExN2n6E4OMtI--t1jRXx6uqzdIV7u5EcLkGQ%3D%3D; _ym_d=1726241603',
            "pragma": "no-cache",
            "priority": "u=1, i",
            # 'referer': 'https://checko.ru/company/pouboetsoogodosaaifr-1022800524042',
            "sec-ch-ua": '"Chromium";v="128", "Not;A=Brand";v="24", "Google Chrome";v="128"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Linux"',
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "user-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36",
        }

        async with self._client.get(url, headers=hdrs, timeout=10, proxy=proxy) as resp:
            raw = await resp.read()
            ...
            try:
                response = await resp.json(content_type=resp.content_type)
            except json.JSONDecodeError:
                ...
                return

        content = response[0]["content"]
        soup = bs4.BeautifulSoup(content, "lxml")
        tag = soup.find("a")
        link = tag["href"]
        return self._base_url + link

    async def get_org_contacts(self, url: str, proxy: str = None):
        url += "?extra=contacts"
        async with self._client.get(url, proxy=proxy, timeout=10) as resp:
            if not resp.ok:
                raise ValueError(f"{resp.reason}, {resp.status}")
            response = await resp.text()

        data = {}

        soup = bs4.BeautifulSoup(response, "lxml")
        contacts_section = soup.find("section", class_="extra-section")

        if "нет сведений" in contacts_section.text.lower():
            return {
                "address": None,
                "phones": None,
                "email": None,
                "website": None,
                "social_networks": None,
                "leader_ein": None,
                "yandex_map_url": None,
            }

        try:
            data["address"] = contacts_section.find_all_next("div")[1].text
        except Exception:
            data["address"] = None

        try:
            data["yandex_map_url"] = contacts_section.find_next("a")["href"]
        except Exception:
            data["yandex_map_url"] = None

        try:
            data["phones"] = [
                t.text
                for t in contacts_section.find(
                    "strong", string=re.compile("Телефон")
                ).find_next_siblings("a")
            ]
        except Exception:
            data["phones"] = []

        try:
            data["email"] = list(
                contacts_section.find("div", string="Веб-сайт")
                .find_parents("div")[0]
                .stripped_strings
            )[-3]
        except Exception:
            data["email"] = []
        try:
            data["website"] = list(
                contacts_section.find("div", string="Веб-сайт")
                .find_parents("div")[0]
                .stripped_strings
            )[-1]
        except Exception:
            data["website"] = []

        try:
            data["social_networks"] = (
                contacts_section.find("div", string="Cоциальные сети")
                .find_next_sibling()
                .text
            )
        except Exception:
            data["social_networks"] = None

        data["leader_ein"] = None
        try:
            async with self._client.get(
                url.rstrip("?extra=contacts"), proxy=proxy, headers=self._hdrs
            ) as resp:
                response = await resp.text()

            soup = bs4.BeautifulSoup(response, "lxml")
            tag = soup.find("span", id=re.compile("copy-leader-inn"))
            data["leader_ein"] = tag.get_text(strip=True)
        except Exception as e:
            ...

        return data


class ListOrgAPI:
    def __init__(self) -> None:
        self._client = aiohttp.ClientSession()
        self._base_url = "https://www.list-org.com"

    @property
    def _hdrs(self):
        headers = {
            "Referer": "https://www.list-org.com/",
            "Upgrade-Insecure-Requests": "1",
        }

        for k, v in ua_generator.generate().headers.get().items():
            headers[k] = v

        return headers

    async def _search_orgs(self, query: str, proxy: str = None):
        url = f"{self._base_url}/search"
        headers = self._hdrs
        params = {
            "val": query,
            "type": "name",
            "work": "on",
            # "staff_min": "1",
            "sort": "",
        }
        headers["Referer"] = f"{self._base_url}/search?val={query}&type=all&sort="

        async with self._client.get(
            url, proxy=proxy, headers=headers, params=params
        ) as resp:
            if not resp.ok:
                raise ValueError(f"{resp.reason}, {resp.status}")
            response = await resp.text()

        soup = bs4.BeautifulSoup(response, "lxml")
        tag = soup.select_one("body > div.main > div.content > p")
        count = int(re.search(r"\d+", tag.text).group())
        return {
            "count": count,
            "pages": int(count // 50 + 1) if count % 50 else count / 50,
        }

    async def _get_orgs_links(self, query: str, page: int, proxy: str = None):
        url = f"{self._base_url}/search"
        headers = self._hdrs
        params = {
            "val": query,
            "type": "all",
            "work": "on",
            # "staff_min": "1",
            "sort": "",
            "page": page,
        }
        headers["Referer"] = f"{self._base_url}/search?val={query}&type=all&sort="

        data = []

        async with self._client.get(
            url, proxy=proxy, headers=headers, params=params
        ) as resp:
            if not resp.ok:
                raise ValueError(f"{resp.reason}, {resp.status}")

            response = await resp.text()

        def collect_with_soup():
            nonlocal data

            soup = bs4.BeautifulSoup(response, "lxml")
            tags_orgs = soup.select("div.card > div > p")
            for tag in tags_orgs:
                tag: bs4.Tag

                info = tag.find("span")
                contents = info.contents

                creds = re.search(r"(\d+).*(\d+)", contents[3])

                obj = {
                    "link": f"{self._base_url}{tag.find('a')['href']}",
                    "full_name": contents[0],
                    "ein": creds.group(1) if creds else None,
                    "kpp": creds.group(2) if creds else None,
                    "address": contents[-1].lstrip(": "),
                }
                data.append(obj)

        res = await asyncio.to_thread(collect_with_soup)
        return data, page

    async def extract_org_data(self, link: str, proxy: str = None):
        headers = self._hdrs
        async with self._client.get(
            link, proxy=proxy, headers=headers, ssl=False, timeout=10
        ) as resp:
            if not resp.ok:
                return
                raise ValueError(f"{resp.reason}, {resp.status}")
            response = await resp.text()

        def extract_with_soup():
            soup = bs4.BeautifulSoup(response, "lxml")

            summary = soup.select_one("table.table-sm")
            rows = summary.find_all("tr")

            full_name = rows[0].find_all("td")[1].text
            leader = rows[1].find_all("td")[1].text

            ein = rows[2].text.split(":")[-1].split("/")[0].strip()
            kpp = rows[2].text.split(":")[-1].split("/")[1].strip()
            personals = next(
                (
                    r.find_all("td")[-1].text
                    for r in rows
                    if r.text.count("Численность персонала:")
                ),
                None,
            )
            status = next(
                (r.find_all("td")[-1].text for r in rows if r.text.count("Статус:")),
                None,
            )

            contact_section = next(
                (
                    s
                    for s in soup.select("div.card.w-100.p-1")
                    if s.text.count("Контактная информация")
                ),
                None,
            )
            contact_rows = contact_section.find_all("p")

            phones = [
                el.text
                for r in contact_rows
                if r.text.count("Телефон:")
                for el in r.find_all("a")
            ]

            emails = [
                el.text
                for r in contact_rows
                if r.text.count("E-mail:")
                for el in r.find_all("a")
            ]

            websites = [
                el.text
                for r in contact_rows
                if r.text.count("Сайт:")
                for el in r.find_all("a")
            ]

            address = [
                r.find("span").text for r in contact_rows if r.text.lower().count("адрес")
            ]

            coords = [
                el.text
                for r in contact_rows
                if r.text.lower().count("координаты")
                for el in r.find_all("a")
            ]

            obj = {
                "full_name": full_name,
                "leader": leader,
                "ein": ein,
                "kpp": kpp,
                "personals": personals,
                "status": status,
                "link": link,
                "phones": phones,
                "emails": emails,
                "websites": websites,
                "address": address,
                "coords": coords,
            }
            return obj

        r = await asyncio.to_thread(extract_with_soup)
        return r

    async def collect_orgs(self, query: str):
        import pickle

        proxy = "http://jMMu8DGnWeMd:RNW78Fm5@pool.proxy.market:10057"
        proxy = "http://90UmMBJx9y9s:RNW78Fm5@185.162.130.85:10543"
        # count = await self._search_orgs(query, proxy=proxy)
        # tasks = [
        #     self._get_orgs_links(query, page, proxy=proxy)
        #     for page in range(1, 32 + 1)
        # ]

        # results = await asyncio.gather(*tasks, return_exceptions=True)

        # data = [
        #     {"data": result[0], "page": result[1]}
        #     for result in results
        #     if not isinstance(result, Exception)
        # ]
        # orgs = [i for o in data for i in o["data"]]
        orgs = pickle.load(open("resp-list-org.pkl", "rb"))
        # data = json.load(open("list-org-base.json"))

        file_data = json.loads(open("list-org-gathered2.json").read())
        target_file_data = json.loads(
            open("data/Юридические-Организации-list_org.json").read()
        )
        file_data = json.loads(open("list-org-base.json").read())

        file_links = [o["link"] for o in target_file_data]
        tasks = []
        needed_links = []

        for item in orgs:
            if item["link"] in file_links:
                continue

            if len(tasks) >= 150:
                data = await asyncio.gather(*tasks, return_exceptions=True)

                r = [d for d in data if isinstance(d, dict)]
                file_data.extend(r)
                json.dump(
                    file_data,
                    open("list-org-base.json", "w"),
                    ensure_ascii=False,
                    indent=2,
                )

                tasks = []
                await asyncio.sleep(10)

            tasks.append(self.extract_org_data(item["link"], proxy))

        data = await asyncio.gather(*tasks, return_exceptions=True)
        r = [d for d in data if isinstance(d, dict)]
        file_data.extend(r)
        json.dump(
            file_data, open("list-org-base.json", "w"), ensure_ascii=False, indent=2
        )
        return data


class FNS:
    def __init__(self) -> None:
        self._client = aiohttp.ClientSession()
        self._base_url = "https://egrul.nalog.ru/"

    @property
    def _hdrs(self):
        headers = {
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7,my;q=0.6",
            "Connection": "keep-alive",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Origin": "https://egrul.nalog.ru",
            "Referer": "https://egrul.nalog.ru/index.html",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
            "X-Requested-With": "XMLHttpRequest",
        }

        for k, v in ua_generator.generate().headers.get().items():
            headers[k] = v

        return headers

    async def _get_items_count(self, query: str) -> dict[str, int]:
        url = self._base_url

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=False)
            page = await browser.new_page()
            await page.goto(url)

            await page.wait_for_selector("#query", state="visible")
            await page.type("#query", query, delay=100)
            await page.click("#btnSearch")

            selector = "#resultPanel"

            await page.wait_for_selector(selector, state="visible")
            content = await page.content()

        soup = bs4.BeautifulSoup(content, "lxml")
        section = soup.select_one(selector)
        tag = section.find("div", class_="result-pager")
        li_tags = tag.find_all("li")
        pages = int(li_tags[-2].text)
        return {"pages": pages, "items": pages * 20}

    async def _get_t_field(self, query: str, page: int = None, proxy: str = None):
        page = page or ""

        url = self._base_url
        data = {
            "vyp3CaptchaToken": "",
            "page": f"{page}",
            "query": query,
            "region": "",
            "PreventChromeAutocomplete": "",
        }

        async with self._client.post(
            url, data=data, headers=self._hdrs, proxy=proxy
        ) as resp:
            response = await resp.json()

        if response.get("ERRORS"):
            raise ValueError(response["ERRORS"])
        return response["t"]

    async def _query_page(self, t_field: str, proxy: str = None):
        headers = {
            "Accept": "*/*",
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7,my;q=0.6",
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "Pragma": "no-cache",
            "Referer": "https://egrul.nalog.ru/index.html",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
            "X-Requested-With": "XMLHttpRequest",
        }

        for k, v in ua_generator.generate().headers.get().items():
            headers[k] = v

        ts = int(datetime.now().timestamp())
        params = {
            "r": ts,
            "_": ts,
        }

        url = f"https://egrul.nalog.ru/search-result/{t_field}?r={ts}&_={ts}"

        async with self._client.get(url, headers=headers, proxy=proxy) as resp:
            response = await resp.json()

        return response

    async def get_organisations_data(self, query: str):
        from loguru import logger

        async def send_packet_query(page: int):
            await asyncio.sleep(random.uniform(0.5, 1.5))

            proxy = f"http://90UmMBJx9y9s:RNW78Fm5@185.162.130.85:{random.randint(10000, 10999)}"
            proxy = None

            t_field = await self._get_t_field(query, page, proxy=proxy)
            await asyncio.sleep(random.uniform(0.5, 1.5))

            data = await self._query_page(t_field, proxy=proxy)
            data["page"] = page
            return data

        # count = await self._get_items_count(query)
        # tasks = [send_packet_query(page) for page in range(1, count["pages"] + 1)]
        tasks = [send_packet_query(page) for page in range(1, 156 + 1)]

        datas = []
        for task in tasks:
            try:
                data = await task
                if not data:
                    continue
            except Exception as e:
                logger.exception(e)
                continue

            print(data["page"])
            datas.append(data)
        return datas


class GeoCoder:
    def __init__(self) -> None:
        self._client = aiohttp.ClientSession()

    async def get_coords(self, address: str, proxy: str = None) -> tuple[float, float]:
        coder = gp.Nominatim(
            user_agent=f"app/{random.randint(1, 10)}",
            adapter_factory=gp.adapters.AioHTTPAdapter,
            proxies={"https": proxy},
            timeout=10,
        )
        async with coder:
            res = await coder.geocode(address, geometry="geojson", timeout=10)

        return res.point.latitude, res.point.longitude

    async def get_geojson(self, address: str, proxy: str = None) -> dict:
        coder = gp.Nominatim(
            user_agent=f"app/{random.randint(1, 10)}",
            adapter_factory=gp.adapters.AioHTTPAdapter,
            proxies={"https": proxy},
            timeout=10,
        )
        async with coder:
            res = await coder.geocode(address, geometry="geojson", timeout=10)

        geojson = res.raw["geojson"]
        return geojson

    async def _query(self, address: str, proxy: str = None):
        headers = {
            "accept": "application/json, text/plain, */*",
            "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7,my;q=0.6",
            "cache-control": "no-cache",
            "origin": "https://apidocs.geoapify.com",
            "pragma": "no-cache",
            "priority": "u=1, i",
            "referer": "https://apidocs.geoapify.com/",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-site",
        }
        for k, v in ua_generator.generate().headers.get().items():
            headers[k] = v

        url = "https://api.geoapify.com/v1/geocode/search"

        replacements = {
            "г. о.": "",
            "город": "",
            "г.": "",
            "д.": "",
            "пр-кт": "проспект",
            "ул ": " ",
            "г ": " ",
        }

        new_address = re.sub(r"^\d+[^,]*,", "", address).strip()
        new_address = re.sub(r",*[^,]+область[^,]*,", "", new_address.lower()).strip()
        new_address = re.sub(r",*[^,]*г\.[^\.]*о\.", "", new_address.lower()).strip()

        for k, v in replacements.items():
            if k in new_address.lower():
                new_address = new_address.replace(k, v)

        new_address = (
            " ".join(dict.fromkeys(new_address.split()).keys()).strip().capitalize()
        )

        params = {
            "text": new_address,
            "format": "json",
            "apiKey": "b8568cb9afc64fad861a69edbddb2658",
            "lang": "ru",
            "filter": "countrycode:ru",
        }

        async with self._client.get(
            url, headers=headers, params=params, proxy=proxy, timeout=10
        ) as resp:
            response = await resp.json()

        return response

    async def geocoder_api_get_coords(self, address: str, proxy: str = None):
        response = await self._query(address, proxy)

        try:
            data = response["results"][0]
            return f"{data['lat']}, {data['lon']}"
        except Exception:
            return None

    async def geocoder_api_get_formatted_address(self, address: str, proxy: str = None):
        response = await self._query(address, proxy)

        try:
            data = response["results"][0]
            return data["formatted"]
        except Exception:
            return None


from playwright.async_api import async_playwright


class GeocoderXYZ:
    def __init__(self) -> None:
        self._client = None
        self._proxy_data = {}

    async def __aenter__(self):
        self._plw = await async_playwright().start()
        self._browser = await self._plw.chromium.launch(headless=True)
        return self

    async def __aexit__(self, exc_type, exc_value, traceback):
        try:
            await self._browser.close()
        except Exception:
            pass

        try:
            await self._plw.stop()
        except Exception:
            pass

    async def get_coords(self, address: str, proxy: str = None):
        if proxy not in self._proxy_data:
            self._proxy_data[proxy] = {"cookies": None, "headers": None}

        url = "https://geocode.xyz/RU"
        headers = {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7,my;q=0.6",
            "cache-control": "no-cache",
            "pragma": "no-cache",
            "priority": "u=0, i",
            "sec-ch-ua": '"Chromium";v="128", "Not;A=Brand";v="24", "Google Chrome";v="128"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Linux"',
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "none",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
        }

        cookies = {}

        if (
            self._proxy_data.get(proxy)
            and self._proxy_data[proxy].get("cookies") is None
        ):
            if proxy:
                proxy_data = URL(proxy)

                proxy_ = {
                    "server": proxy,
                    "username": proxy_data.user,
                    "password": proxy_data.password,
                }
            else:
                proxy_ = {}

            context = await self._browser.new_context(proxy=proxy_)
            page = await context.new_page()

            await page.goto(url)

            await page.wait_for_selector("#hello", state="visible")
            user_agent_pw = await page.evaluate("navigator.userAgent")
            sec_ua_pw = await page.evaluate("""
                () => {
                    return {
                        'sec-ch-ua': navigator.userAgentData.brands.map(b => `${b.brand};v="${b.version}"`).join(', '),
                        'sec-ch-ua-mobile': navigator.userAgentData.mobile ? '?1' : '?0',
                        'sec-ch-ua-platform': navigator.userAgentData.platform
                    };
                }
            """)

            headers["user-agent"] = user_agent_pw

            cookies_pw = await context.cookies(url)
            cookies = {row["name"]: row["value"] for row in cookies_pw}

        new_headers = {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7,my;q=0.6",
            "cache-control": "no-cache",
            "origin": "https://geocode.xyz",
            "pragma": "no-cache",
            "priority": "u=0, i",
            "referer": "https://geocode.xyz/RU",
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "same-origin",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
            "user-agent": user_agent_pw,
        }
        new_headers.update(sec_ua_pw)
        url = "https://geocode.xyz"
        data = {
            "locate": address,
            "geoit": "JSON",
            "region": "RU",
            "ok": "Geocode",
        }

        if (
            self._proxy_data.get(proxy)
            and self._proxy_data[proxy].get("cookies") is None
        ):
            self._proxy_data[proxy]["cookies"] = cookies
            self._proxy_data[proxy]["headers"] = new_headers
        else:
            cookies = self._proxy_data[proxy]["cookies"]
            new_headers = self._proxy_data[proxy]["headers"]

        async with aiohttp.ClientSession() as session:
            form = aiohttp.FormData()
            for k, v in data.items():
                form.add_field(k, v)

            async with session.post(
                url,
                cookies=cookies,
                headers=new_headers,
                proxy=proxy,
                timeout=20,
                data=form,
                ssl=False,
            ) as resp:
                response = await resp.json()

        return response["latt"], response["longt"]


async def gather_polygons_dosaaf():
    data = await parse_regional_departments()
    data_geocoded = []
    tasks = []

    async def geocode(query: str, fd: str = None):
        nonlocal data_geocoded

        # proxy = f"http://90UmMBJx9y9s:RNW78Fm5@185.162.130.85:{random.randint(10000, 10999)}"
        proxy = None
        
        inst = GeoCoder()
        # address = await inst.geocoder_api_get_formatted_address(query, proxy)
        address = query
        geojson = await inst.get_geojson(query, proxy)
        await inst._client.close()
        data_geocoded.append({"name": query, "address": address, "geojson": geojson, "federal_district": fd})

    while True:
        file_data = orjson.loads(open("regions-geocoded.json").read())
        existing_names = [r["name"] for r in file_data if r.get("name")]
        for fd in data:
            fd_name = fd["name"]

            if fd_name not in existing_names:
                task = geocode(fd_name)
                tasks.append(task)
                print(f"fd: {fd_name}")

            for region in fd["regions"]:
                region_name = region["name"]

                if region_name not in existing_names:
                    if region_name == "Республика Северная Осетия":
                        region_name = "Северная Осетия"
                    
                    print(f"region: {region_name}")
                    task = geocode(region_name, fd_name)
                    tasks.append(task)

            if len(tasks) > 10:
                res = await asyncio.gather(*tasks, return_exceptions=True)
                tasks.clear()

                exists = orjson.loads(open("regions-geocoded.json").read())
                exists.extend(data_geocoded)
                tmp = []
                [tmp.append(d) for d in exists if d["name"] not in [o["name"] for o in tmp]]
                open("regions-geocoded.json", "w").write(orjson.dumps(tmp).decode("utf-8"))

                data_geocoded.clear()

        if not tasks:
            break

        res = await asyncio.gather(*tasks, return_exceptions=True)
        if data_geocoded:
            exists = orjson.loads(open("regions-geocoded.json").read())
            exists.extend(data_geocoded)

            tmp = []
            [tmp.append(d) for d in exists if d["name"] not in [o["name"] for o in tmp]]
            open("regions-geocoded.json", "w").write(orjson.dumps(tmp).decode("utf-8"))
            data_geocoded.clear()
            tasks.clear()

        await asyncio.sleep(2)


async def gather_military_districts():
    names = {
        "Московский военный округ": [
            "Белгородская область",
            "Брянская область",
            "Владимирская область",
            "Воронежская область",
            "Ивановская область",
            "Калужская область",
            "Костромская область",
            "Курская область",
            "Липецкой области",
            "Москвы",
            "Московская область",
            "Нижегородская область",
            "Орловская область",
            "Рязанской области",
            "Смоленская область",
            "Тамбовская область",
            "Тверская область",
            "Тульская область",
            "Ярославская область",
        ],
        "Ленинградский военный округ": [
            "Архангельская область",
            "Вологодская область",
            "Калининградская область",
            "Карелия",
            "Коми",
            "Ленинградская область",
            "Мурманская область",
            "Ненецкий автономный округ",
            "Псковская область",
            "Санкт-Петербург",
        ],
        "Южный военный округ": [
            "Республика Адыгея",
            "Республика Дагестан",
            "Республика Ингушетия",
            "Кабардино-Балкарская Республика",
            "Республика Калмыкия",
            "Карачаево-Черкесская Республика",
            "Республика Северная Осетия — Алания",
            "Чеченская Республика",
            "Краснодарский край",
            "Ставропольский край",
            "Астраханская область",
            "Волгоградская область",
            "Ростовская область",
        ],
        "Центральный военный округ": [
            "Ханты-Мансийский автономный округ - Югра",
            "Ямало-Ненецкий автономный округ",
            "Республика Алтай",
            "Республика Башкортостан",
            "Республика Марий Эл",
            "Республика Мордовия",
            "Республика Татарстан",
            "Республика Тыва",
            "Удмуртская Республика",
            "Республика Хакасия",
            "Чувашская Республика",
            "Иркутская область",
            "Кемеровская область",
            "Кировская область",
            "Курганская область",
            "Новосибирская область",
            "Омская область",
            "Оренбургская область",
            "Пензенская область",
            "Самарская область",
            "Саратовская область",
            "Алтайский край",
            "Красноярский край",
            "Пермский край",
        ],
        "Восточный военный округ": [
            "Республики Бурятия",
            "Республики Саха (Якутия)",
            "Забайкальский край",
            "Камчатский край",
            "Приморский край",
            "Хабаровский край",
            "Амурская область",
            "Магаданская область",
            "Сахалинская область",
            "Еврейская автономная область",
            "Чукотского автономного округа",
        ],
    }

    data_geocoded = {}

    async def geocode(name: str, district_name: str):
        nonlocal data_geocoded

        proxy = f"http://90UmMBJx9y9s:RNW78Fm5@185.162.130.85:{random.randint(10000, 10999)}"
        proxy = None
        inst = GeoCoder()
        # address = await inst.geocoder_api_get_formatted_address(name, proxy)
        address = name
        geojson = await inst.get_geojson(address, proxy)
        await inst._client.close()

        if data_geocoded.get(district_name) is None:
            data_geocoded[district_name] = []

        data_geocoded[district_name].append(
            {
                "name": name,
                "address": address,
                "geojson": geojson,
            }
        )

    while True:
        tasks = []
        # load existing data from file json
        existsing_data = orjson.loads(open("military-districts.json").read())
        # get all names from file data
        existsing_names = [
            o["name"]
            for district_name in existsing_data
            for o in existsing_data[district_name]
        ]

        if all(name in existsing_names for name in names.values()):
            break

        for district_name in names:
            part = [
                geocode(name, district_name)
                for name in names[district_name]
                if name not in existsing_names
            ]
            if not part:
                continue
            tasks.extend(part)

            # if len(tasks) > 10:
            res = await asyncio.gather(*tasks, return_exceptions=True)

            tmp = existsing_data.copy()
            if district_name not in tmp:
                tmp[district_name] = []

            [
                tmp[district_name].append(d)
                for d in data_geocoded[district_name]
                if d not in tmp[district_name]
            ]

            ...
            open("military-districts.json", "w").write(
                orjson.dumps(tmp).decode("utf-8")
            )
            data_geocoded.clear()
            tasks.clear()

        if tasks:
            raise NotImplementedError
            res = await asyncio.gather(*tasks, return_exceptions=True)

            tmp = {}
            tmp[district_name] = []
            [tmp[district_name].append(d) for d in data_geocoded[district_name]]

            ...
            open("military-districts.json", "w").write(
                orjson.dumps((existsing_data | tmp)).decode("utf-8")
            )
            data_geocoded.clear()
            tasks.clear()


class ReverseGeocoder:
    async def get_data_by_coords(self, coords: tuple[float, float], proxy: str = None):
        coder = gp.Nominatim(
            user_agent=f"app/{random.randint(1, 10)}",
            adapter_factory=gp.adapters.AioHTTPAdapter,
            proxies={"https": proxy},
            timeout=10,
        )
        async with coder:
            res = await coder.reverse(coords, timeout=10, language="ru")

        return res.raw


async def reverse_geocoding():
    coder = ReverseGeocoder()
    df = pd.read_json("data/Юридические-Организации-list_org.json")

    def get_federal_district(data: dict):
        return data["address"]["region"]

    def get_region_name(data: dict):
        return data["address"]["state"]

    proxy = (
        f"http://90UmMBJx9y9s:RNW78Fm5@185.162.130.85:{random.randint(10000, 10999)}"
    )
    data_geocoded = []

    async def get_data(row):
        nonlocal data_geocoded

        coords = row["coords"][0]
        try:
            data = await coder.get_data_by_coords(coords)
        except Exception as e:
            data = await coder.get_data_by_coords(coords, proxy)

        levels = {
            "1": data["address"]["country"],
            "2": get_federal_district(data),
            "3": get_region_name(data),
            "4": data["display_name"],
        }

        obj = row.to_dict()
        obj["geolevels"] = levels

        data_geocoded.append(obj)

    tasks = []
    n = 15
    already_geocoded = orjson.loads(open("reverse-geocoded.json").read())
    already_eins = [str(r["ein"]) for r in already_geocoded]
    for i, row in df.iterrows():
        if str(row["ein"]) in already_eins:
            continue

        if len(tasks) > n:
            res = await asyncio.gather(*tasks, return_exceptions=True)
            tasks.clear()
            await asyncio.sleep(5)

        task = get_data(row)
        tasks.append(task)
        ...
    res = await asyncio.gather(*tasks, return_exceptions=True)
    ...
    already_geocoded.extend(data_geocoded)
    open("reverse-geocoded.json", "w").write(
        orjson.dumps(already_geocoded).decode("utf-8")
    )


async def parse_ec():
    wb = load_workbook("data/ЕЦ.xlsx", data_only=True)
    needed_shetnames = {
        "центральный": "Центральный федеральный округ",
        "северо-западный": "Северо-западный федеральный округ",
        "сибирский": "Сибирский федеральный округ",
        "приволжский": "Приволжский федеральный округ",
        "уральский": "Уральский федеральный округ",
        "северо-кавказский": "Северо-кавказский федеральный округ",
        "дальневосточный": "Дальневосточный федеральный округ",
        "юфо": "Южный федеральный округ",
    }
    data = {}

    for ws in wb.worksheets:
        if ws.title.lower().strip() not in needed_shetnames:
            continue

        fd = needed_shetnames[ws.title.lower().strip()]
        rc = None
        print(fd)
        search_childs_ec = False

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cells = [cell for cell in row]
            cells_values = [c.value for c in cells]
            if all(v is None for v in cells_values):
                continue

            if search_childs_ec is False:
                if all(not isinstance(c, MergedCell) for c in cells[2:6]):
                    continue

                ec_cell = cells[0] if cells_values[0] else cells[1]
                if not ec_cell.value or ec_cell.fill.patternType != "solid":
                    continue

                # Means the cell is not needed
                if ec_cell.fill.fgColor.rgb == "FFFFFF00":
                    continue
                if ec_cell.value.lower().count("федеральный округ") > 0:
                    continue

                formatted_name = " ".join(
                    [
                        w.strip()
                        for w in re.sub(r"\d*(\(.+\))*", "", ec_cell.value)
                        .strip()
                        .strip(".")
                        .split()
                        if w.strip()
                    ]
                ).strip()

                search_childs_ec = True
                print(formatted_name)
                
                rc = re.sub(r" г\.|ЕМЦ|[\d\.]|\(.*\)", "", formatted_name).strip()
                data[formatted_name] = []
            elif any(c.fill.fgColor.rgb == "FFFFFF00" for c in cells[0:3]) and any(
                c.value.lower().count("итого") for c in cells[0:3] if c.value
            ):
                search_childs_ec = False
                # print("\n\n")
                continue
            else:
                object_name = cells_values[1]
                place = cells_values[2]
                formatted_place = None

                for k, v in abbreviations.items():
                    if isinstance(v, list):
                        v = v[-1]

                    if k in [w.strip() for w in object_name.split()]:
                        object_name = object_name.replace(k, v)
                    if (
                        not formatted_place
                        and place
                        and k in [w.strip() for w in place.split()]
                    ):
                        chunk_words = []
                        for w in place.split():
                            w = w.strip()
                            if w == k:
                                w = v
                            chunk_words.append(w)
                        formatted_place = " ".join(chunk_words)

                if place:
                    place = " ".join([w.strip() for w in place.split()])
                if formatted_place:
                    formatted_place = " ".join(
                        [w.strip() for w in formatted_place.split()]
                    )
                if object_name:
                    object_name = " ".join([w.strip() for w in object_name.split()])

                data[formatted_name].append(
                    {
                        "name": object_name,
                        "place": place,
                        "formatted_place": formatted_place,
                        "region": rc,
                        "federal_district": fd,
                    }
                )
                ...
    open("ecs.json", "w").write(orjson.dumps(data).decode("utf-8"))


async def concat_orgs_and_ec():
    # orgs_objs = json.load(open("data/Юридические-Организации-list_org.json"))
    orgs_objs = json.load(open("list-org-base.json"))
    ecs = json.load(open("data/Единые-центры.json"))
    orgs_names = [o["full_name"] for o in orgs_objs]

    def parse_city(incorrect_word: str):
        patterns = {
            r"(.*)г\.([ ]+.*)": r"\1г. \2",
            r".+ центр": "",
            r"(ЮВАО|ЮАО|САО) г\. (.*)": r"\1 город \2",
            r"[^(ЮВАО|ЮАО|САО)]+г\.+(.+)": r"\1",
            r".+аэр\. (\w+)": r"аэропорт \1",
            r"[Мм]естное отделение": "",
            r"ЗЦ|РЦ|МЗЦ": "",
            r"\"(г\.|город.+)\"": r"\1",
            r"[Оо]бл.( |$)": "область",
            r"[«»,\(\)\"\']": "",
            # r"-": " ",
            r"КБР": "",
            r"с\.": "",
            r"област\w+|район\w*|станица|г\.|город\w+": "",
            # r"я( |$)": " ",
            r"ой|ом[$\s]|ому|ий|ый": "",
            r"Севастопольск АСК": "",
            "Севастополя": "Севастопол",
            "Ю ": " ",
            "ВКерчь": "Керчь",
            "области": "",
            r"ая[$\s]": "",
        }

        name = " ".join([w.strip() for w in incorrect_word.split()])

        for pattern in patterns:
            replacement = patterns[pattern]
            if name.lower().count('белгород') and pattern.count('город'):
                continue
                
            name = re.sub(pattern, replacement, name)
            ...

        for abbr in abbreviations:
            name = name.replace(abbr, "")

        if not name:
            ...

        name = " ".join([w.strip() for w in name.split()]).strip()
        name = name.replace("ё", "е").replace("Ё", "Е")
        # print(name)
        return name

    address_patterns = {
        r".+центр\s{0,2}": "",
        r"ЗЦ|РЦ": "",
        # r"[\(\)]": r"",
        # r"\s{0,2}центр": r"",
    }
    clean_pattern = re.compile(r"\"',\.")
    
    def get_org_types(name: str):
        correct_word = " ".join([w.replace("-", " ").strip() for w in name.split()])
        types = []
        for key, value in abbreviations.items():
            if isinstance(value, str):
                value = [value.lower()]
            
            value = [re.sub(r"\s{2,}", " ", v.lower().replace("-", " ").strip()) for v in value]
            ...
            types.extend([v.lower() for v in value if correct_word.lower().count(v.lower())])

        return list(set(types))


    data = []
    items_objs = []
    tmp = []
    all_ = 0
    items = []
    not_found = 0
    for ec_name in ecs:
        if not ec_name.startswith("ЕМЦ "):
            continue

        ec_location = ec_name.split("ЕМЦ ")[-1]
        break_ = False

        for item in ecs[ec_name]:
            if item["name"] in items:
                continue
            items.append(item["name"])
            all_ += 1
            item["related_org"] = None
            
            for _n in [item["name"], ec_name]:    
                find = False
                name = re.sub(clean_pattern, "", _n)
                name = name.replace("ё", "е").replace("Ё", "Е")

                place = item["formatted_place"]
                if not place:
                    break_ = True
                    break

                name = parse_city(name)
                item_location = name
                if item_location.startswith('Е '):
                    item_location = item_location.replace('Е ', '')
                if item_location.lower().count('область'):
                    item_location = item_location.replace('область', '')
                item_location = parse_city(item_location)
                    
                item_location = " ".join([w.strip() for w in item_location.split()])

                for org in orgs_objs:
                    org_name = " ".join([w.strip() for w in org["full_name"].split()])
                    org_types = get_org_types(org_name)
                    
                    org_types = [
                        " ".join([w.strip() for w in t.split()]).lower() for t in org_types
                    ]
                    org_location = org["address"]
                    
                    if not org_name.lower().count(item_location.lower()):
                        continue
                    
                    for org_type in org_types:
                        if org_type == 'автомобильная школа' and 'школа' not in org_types:
                            org_types.append('школа')

                    if not org_types:
                        continue

                    if org_location:
                        org_location = org_location[0]
                        
                    
                    # if (
                    #     org_location and org_location.lower().count(item_location.lower())
                    # ):
                    if any(place.lower().strip().count(t.lower().strip()) or t.lower().strip().count(place.lower().strip()) for t in org_types):
                        find = True
                        item["related_org"] = org["ein"]
                        print(f"{name} - {org['address']}")
                        break_ = True
                        break
                    else:
                        if item not in tmp:
                            # print(f"[{item_location}] - [{org_name}] - {item}\n")
                            tmp.append(item)
                        ...    
                
                if find:
                    break
                
            items_objs.append(item)
            
            if break_:
                continue
            
            if not find:
                # print(f"{item["name"]} - {item_location} - {place}")
                not_found += 1
                continue

    print(f"Not found: {not_found}/{all_}")
    print(len(tmp))
    
    json.dump(items_objs, open("items_objs.json", "w"), indent=2, ensure_ascii=False)



async def format_data():
    rows = []
    
    async def geocode(row: pd.Series):
        nonlocal rows
        inst = GeoCoder()
        listorg = ListOrgAPI()
        proxy = f"http://90UmMBJx9y9s:RNW78Fm5@185.162.130.85:{random.randint(10000, 10999)}"
        
        obj = row.to_dict()
        
        link = obj["link"]
        _obj = await listorg.extract_org_data(link, proxy)
        if _obj:
            obj.update(_obj)
        
        addr = obj["address"]
        ein = obj["ein"]
        
        if isinstance(addr, list):
            if not addr:
                return
            addr = addr[0]

        addr = await inst.geocoder_api_get_formatted_address(addr, proxy=proxy)
        if not addr:
            raise ValueError(f"No address for {ein}")

        if isinstance(addr, list):
            addr = addr[0]

        coords = await inst.geocoder_api_get_coords(addr, proxy=proxy)
        if not coords:
            raise ValueError(f"No coords for {ein}")
        
        coder = ReverseGeocoder()
        try:
            data = await coder.get_data_by_coords(coords)
        except Exception as e:
            data = await coder.get_data_by_coords(coords, proxy)
            
        def get_federal_district(data: dict):
            return data["address"]["region"]

        def get_region_name(data: dict):
            return data["address"]["state"]
        
        geolevels = {
            "1": data["address"]["country"],
            "2": get_federal_district(data),
            "3": get_region_name(data),
            "4": data["display_name"],
        }

        await inst._client.close()

        if not coords:
            coords = []
        else:
            coords = [coords]

        obj["coords"] = coords
        obj["geolevels"] = geolevels
        rows.append(obj)
        
    df = pd.read_json("list-org-base.json")
    
    tasks = []
    for i, row in df.iterrows():
        if isinstance(row["geolevels"], dict) and row["geolevels"].get("4"):
            continue
        
        if isinstance(row["coords"], list) and row["coords"]:
            continue
        
        tasks.append(geocode(row))
        
    res = await asyncio.gather(*tasks, return_exceptions=False)
    
    src = df.fillna("").to_dict('records')
    src2 = src.copy()
    
    for obj in rows:
        srcobj = [o for o in src2 if int(o["ein"]) == int(obj["ein"])][0]
        i = src2.index(srcobj)
        src2[i].update(obj)
        
    json.dump(src2, open("list-org-base.json", "w"), indent=2, ensure_ascii=False)
    return
    
    orgs_objs = json.load(open("list-org-base.json"))
    ecs = json.load(open("data/Единые-центры.json"))
    orgs = []
    
    def parse_city(incorrect_word: str):
        patterns = {
            r"(.*)г\.([ ]+.*)": r"\1г. \2",
            r".+ центр": "",
            r"(ЮВАО|ЮАО|САО) г\. (.*)": r"\1 город \2",
            r"[^(ЮВАО|ЮАО|САО)]+г\.+(.+)": r"\1",
            r".+аэр\. (\w+)": r"аэропорт \1",
            r"[Мм]естное отделение": "",
            r"ЗЦ|РЦ|МЗЦ": "",
            r"\"(г\.|город.+)\"": r"\1",
            r"[Оо]бл.( |$)": "область",
            r"[«»,\(\)\"\']": "",
            # r"-": " ",
            r"КБР": "",
            r"с\.": "",
            r"област\w+|район\w*|станица|г\.|город\w+": "",
            # r"я( |$)": " ",
            r"ой|ом[$\s]|ому|ий|ый": "",
            r"Севастопольск АСК": "",
            "Севастополя": "Севастопол",
            "Ю ": " ",
            "ВКерчь": "Керчь",
            "области": "",
            r"ая[$\s]": "",
        }

        name = " ".join([w.strip() for w in incorrect_word.split()])

        for pattern in patterns:
            replacement = patterns[pattern]
            if name.lower().count('белгород') and pattern.count('город'):
                continue
                
            name = re.sub(pattern, replacement, name)
            ...

        for abbr in abbreviations:
            name = name.replace(abbr, "")

        if not name:
            ...

        name = " ".join([w.strip() for w in name.split()]).strip()
        name = name.replace("ё", "е").replace("Ё", "Е")
        # print(name)
        return name

    clean_pattern = re.compile(r"\"',\.")    
    def get_org_types(name: str):
        correct_word = " ".join([w.replace("-", " ").strip() for w in name.split()])
        types = []
        for key, value in abbreviations.items():
            if isinstance(value, str):
                value = [value.lower()]
            
            value = [re.sub(r"\s{2,}", " ", v.lower().replace("-", " ").strip()) for v in value]
            ...
            types.extend([v.lower() for v in value if correct_word.lower().count(v.lower())])

        return list(set(types))

    items_objs = []
    tmp = []
    all_ = 0
    items = []
    not_found = 0
    for ec_name in ecs:
        if not ec_name.startswith("ЕМЦ "):
            continue

        break_ = False

        for item in ecs[ec_name]:
            if item["name"] in items:
                continue
            items.append(item["name"])
            all_ += 1
            item["related_org"] = None
            
            for _n in [item["name"], ec_name]:    
                find = False
                name = re.sub(clean_pattern, "", _n)
                name = name.replace("ё", "е").replace("Ё", "Е")

                place = item["formatted_place"]
                if not place:
                    break_ = True
                    break

                name = parse_city(name)
                item_location = name
                if item_location.startswith('Е '):
                    item_location = item_location.replace('Е ', '')
                if item_location.lower().count('область'):
                    item_location = item_location.replace('область', '')
                item_location = parse_city(item_location)
                    
                item_location = " ".join([w.strip() for w in item_location.split()])

                for org in orgs_objs:
                    org_name = " ".join([w.strip() for w in org["full_name"].split()])
                    org_types = get_org_types(org_name)
                    
                    org_types = [
                        " ".join([w.strip() for w in t.split()]).lower() for t in org_types
                    ]
                    org_location = org["address"]
                    
                    if not org_name.lower().count(item_location.lower()):
                        continue
                    
                    for org_type in org_types:
                        if org_type == 'автомобильная школа' and 'школа' not in org_types:
                            org_types.append('школа')

                    if not org_types:
                        continue

                    if org_location:
                        org_location = org_location[0]
                        
                    if any(place.lower().strip().count(t.lower().strip()) or t.lower().strip().count(place.lower().strip()) for t in org_types):
                        find = True
                        item["related_org"] = org["ein"]
                        print(f"{name} - {org['address']}")
                        break_ = True
                        
                        federal_district = None
                        region = None
                        
                        if isinstance(org["geolevels"], dict):
                            federal_district = org["geolevels"]["2"]
                            region = org["geolevels"]["3"]
                        
                        
                        type = Organisation(
                            full_name=org["full_name"],
                            state=org["status"],
                            ein=str(org["ein"]),
                            kpp=str(org["kpp"]),
                            contact_emails=org["emails"],
                            contact_phones=org["phones"],
                            websites=org["websites"],
                            type_org=org["type_org"],
                            coords=org["coords"],
                            address=org["address"],
                            federal_district=federal_district,
                            region=region,
                            link_listorg=org["link"],
                            personals=int(org["personals"]) if org["personals"] else None,
                        )
                        
                        if item["name"].lower().count("ец "):
                            type_ec = 'Единый центр'
                        elif item["name"].lower().count("емц "):
                            type_ec = 'Единый межрегиональный центр'
                        elif item["name"].lower().count("зц"):
                            type_ec = 'Зональный центр'
                        elif item["name"].lower().count("рц"):
                            type_ec = 'Региональный центр'
                        elif item["name"].lower().count("региональный единый центр "):
                            type_ec = 'Региональный единый центр'
                        elif item["name"].lower().count("зональный многофункциональный центр "):
                            type_ec = 'Зональный многофункциональный центр'
                        elif item["name"].lower().count("зональный центр "):
                            type_ec = 'Зональный центр'
                        else:
                            raise ValueError(f"Unknown type ec: {item['name']}")
                        
                        type_ec = EC(
                            name=item["name"],
                            region=item["region"],
                            federal_district=item["federal_district"],
                            type_ec=type_ec,
                            organisations=[str(org["ein"])],
                            address=org["address"],
                            coords=org["coords"],
                        )
                        
                        orgs.append(type_ec)
                        break
                    else:
                        if item not in tmp:
                            # print(f"[{item_location}] - [{org_name}] - {item}\n")
                            tmp.append(item)
                        ...    
                
                if find:
                    break
                
            items_objs.append(item)
            
            if break_:
                continue
            
            if not find:
                # print(f"{item["name"]} - {item_location} - {place}")
                not_found += 1
                continue

    print(f"Not found: {not_found}/{all_}")
    print(len(tmp))
    import pickle
    pickle.dump(orgs, open("orgs.pkl", "wb"))
    # json.dump(items_objs, open("items_objs.json", "w"), indent=2, ensure_ascii=False)
    ...



import pickle
async def main():
    await gather_polygons_dosaaf()
    ...
    # await parse_regional_departments()
    # _ = pickle.load(open("orgs-ecs.pkl", "rb"))
    
    # await prepare_db()
    # orgs = _["orgs"]
    # ecs = _["ecs"]
    
    # async with get_sqlalchemy_async_sessionmaker()() as session:
    #     repo_orgs = OrganisationRepository(session)
    #     repo_ecs = ECRepository(session)
        
    #     for org in orgs:
    #         ...
    #         await repo_orgs.add(org)
        
    #     for ec in ecs:
    #         ...
    #         await repo_ecs.add(ec)
        
    return
    return await format_data()
    # await prepare_db()
    
    ecs = json.load(open("data/Единые-центры.json")) 
    df = pd.read_json("list-org-base.json")
    df = df.fillna(np.nan).replace([np.nan], [None])
    df = df.to_dict("records")
    
    orgs = []
    
    for obj in df:
        
        federal_district = None
        region = None
        
        if isinstance(obj["geolevels"], dict):
            federal_district = obj["geolevels"]["2"]
            region = obj["geolevels"]["3"]
        
        type = Organisation(
            full_name=obj["full_name"],
            state=obj["status"],
            ein=str(obj["ein"]),
            kpp=str(obj["kpp"]),
            contact_emails=obj["emails"],
            contact_phones=obj["phones"],
            websites=obj["websites"],
            type_org=obj["type_org"],
            coords=obj["coords"],
            address=obj["address"],
            federal_district=federal_district,
            region=region,
            link_listorg=obj["link"],
            personals=int(obj["personals"]) if obj["personals"] else None,
        )
        orgs.append(type)
        
    for ec_name in ecs:
        for item in ecs[ec_name]:   
            if item["name"].lower().count("ец "):
                type_ec = 'Единый центр'
            elif item["name"].lower().count("емц "):
                type_ec = 'Единый межрегиональный центр'
            elif item["name"].lower().count("зц"):
                type_ec = 'Зональный центр'
            elif item["name"].lower().count("рц"):
                type_ec = 'Региональный центр'
            elif item["name"].lower().count("региональный единый центр "):
                type_ec = 'Региональный единый центр'
            elif item["name"].lower().count("зональный многофункциональный центр "):
                type_ec = 'Зональный многофункциональный центр'
            elif item["name"].lower().count("зональный центр "):
                type_ec = 'Зональный центр'
            else:
                raise ValueError(f"Unknown type ec: {item['name']}")
            
            type = EC(
                name=item["name"],
                region=item["region"],
                federal_district=item["federal_district"],
                type_ec=type_ec,
            )
            
            print(type)
           
        # async with get_sqlalchemy_async_sessionmaker()() as session:
        #     repo = OrganisationRepository(session)
        #     await repo.add(type)
         
    # await format_data()
    # return
    # await concat_orgs_and_ec()
    # await prepare_db()
    # await csv2db()
    return

    # df = json.load(open("data/Федеральные_округа-Регионы.json"))
    # df = json.load(open("data/Юридические-Организации-list_org.json"))
    # exists = []
    # rows = []

    # for i, obj in enumerate(df):
    #     name = obj["full_name"]

    #     type = obj["type_org"]
    #     if type == "Профессиональное образовательное учреждение":
    #         _type = _transform_object_name(name.lower().replace(type.lower(), ""))[1]
    #         if _type:
    #             type = _type

    #     obj["type_org"] = type

    #     if type and type not in exists:
    #         exists.append(type)
    #         # print(type)

    #     rows.append(obj)
    # print(obj["name"])
    # print("\n\n")

    # await parse_ec()
    # return await gather_military_districts()
    # json.dump(rows, open("data/Юридические-Организации-list_org.json", "w"), indent=2, ensure_ascii=False)
    # return
    df_orig = pd.read_csv(
        "/home/daniil/VSCode/dosaaf-projects/utils/tables_parser/organisation-with-geos-full-tmp.csv"
    )
    proxy = (
        f"http://90UmMBJx9y9s:RNW78Fm5@185.162.130.85:{random.randint(10000, 10999)}"
    )
    df2 = pd.read_json("list-org-base.json")
    df = pd.read_json("data/Юридические-Организации-list_org.json")
    checko = CheckoAPI()
    geo = GeocoderXYZ()

    async def geocode(row):
        nonlocal rows
        inst = GeoCoder()

        addr = row["address"]
        ein = row["ein"]

        if not addr:
            # link = await checko._query(ein, proxy)
            # checko_data = await checko.get_org_contacts(link, proxy)
            # addr = checko_data["address"]
            addr = await inst.geocoder_api_get_formatted_address(addr, proxy=proxy)
            # addr = await geo.get_coords(addr, proxy=proxy)

            if not addr:
                raise ValueError(f"No address for {ein}")

        if isinstance(addr, list):
            addr = addr[0]

        # coords = await inst.geocoder_api_get_coords(addr, proxy=proxy)
        coords = await geo.get_coords(addr, proxy=proxy)
        if not coords:
            coords = []

        await inst._client.close()

        if not coords:
            coords = []
        else:
            coords = [coords]

        row["coords"] = coords
        row["address"] = addr
        rows.append(row.to_dict())

    rows = []
    tasks = []
    existing_eins = [str(r["ein"]) for _, r in df.iterrows() if r["ein"]]
    for i, row in df2.iterrows():
        if isinstance(row.to_dict().get("geolevels"), dict):
            continue

        if row["coords"]:
            continue

        if str(row["ein"]) in existing_eins:
            continue

        if len(tasks) > 10:
            async with geo:
                res = await asyncio.gather(*tasks, return_exceptions=True)
                tasks.clear()
                await asyncio.sleep(5)

        tasks.append(geocode(row))
        continue

    res = await asyncio.gather(*tasks, return_exceptions=True)

    # api = ListOrgAPI()
    # await api.collect_orgs("ДОСААФ")
    return
    # types = await get_types()
    # json.dump(types, open("types_orgs.json", "w"), indent=2, ensure_ascii=False)

    # data = await parse_regional_departments()
    data = json.load(open("places.json"))

    another = []
    for obj in data:
        if obj.get("geojson"):
            continue
        coder = gp.Nominatim(user_agent=f"app/{random.randint(1, 10)}")

        fd = obj["name"]
        res = coder.geocode(fd, geometry="geojson")
        ...

        obj["geojson"] = json.dumps(res.raw["geojson"], ensure_ascii=False, indent=2)

        for region in obj["regions"]:
            if region.get("geojson"):
                continue

            res = coder.geocode(f"{fd}, {region["name"]}", geometry="geojson")
            ...
            ...
            region["geojson"] = json.dumps(
                res.raw["geojson"], ensure_ascii=False, indent=2
            )

        another.append(obj)
        await asyncio.sleep(3)
    ...


if __name__ == "__main__":
    import asyncio

    asyncio.run(main())
