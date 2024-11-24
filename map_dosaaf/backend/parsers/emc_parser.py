import re
import openpyxl
import orjson


def load_dataset(file_path: str):
    """Construct a list of dict objects with `едиными центрами` from excel table

    Args:
        file_path (str): path to excel

    Returns:
        list[dict]: list with data about единых центрах
    """
    workbook = openpyxl.load_workbook(file_path)
    # sheet = workbook.active
    data = []
    
    emc = None
    obj = {}
    search_ecs = False
    
    for i, sheet in enumerate(workbook.worksheets):
        if i not in range(2, 9 + 1):
            continue
        
        for row in sheet.iter_rows(min_row=0, max_col=sheet.max_column, max_row=sheet.max_row):
            for cell in row:
                if not cell or not isinstance(cell.value, str):
                    continue
                
                try:
                    if cell.value.lower().count("емц") and cell.fill.bgColor.rgb == 'FFC0C0C0':
                        value = re.sub(r"\(.*", "", cell.value)
                        value = re.sub(r"обл\.", "область", value)
                        value = re.sub(r"г\.", r"город", value)
                        value = re.sub(r"[Рр]есп\.", "Республика", value)
                        value = re.sub(r"\d[\.\s]*", "", value)
                        value = " ".join([w for w in value.split()])
                        emc = value
                except Exception as e:
                    pass
            
            if emc is None:
                for cell in row:
                    try:
                        
                        if not cell or not isinstance(cell.value, str):
                            continue
                        
                        if cell.value.lower().count("итого"):
                            search_ecs = False
                            
                        if not search_ecs:
                            break
                        
                        if not any(cell.value.count(w) for w in ["ЗЦ", "РЦ"]):
                            continue
                        
                        value = cell.value
                        value = " ".join(w.strip() for w in value.split())
                        
                        i = row.index(cell)
                        based_on = row[i+1].value
                        
                        obj["единые центры"].append({
                            "name": value,
                            "based_on": based_on
                        })
                        
                    except Exception:
                        pass
                    
                continue
            
            obj = {
                "емц": emc,
                "единые центры": []
            }
            data.append(obj)
            emc = None
            search_ecs = True
    
    return data


def main():
    from pprint import pprint
    res = load_dataset("data/ЕЦ.xlsx")
    # orgs = orjson.loads('data/Юридические-Организации-list_org.json')
    
    for item in res:
        for ec in item["единые центры"]:
            emc = item["емц"]
            if emc is None:
                continue
            
            value = ec["name"]
            value = re.sub(r"(.*)\((.+)\)", r"\1 \2", value)
            value = re.sub(r"обл\.", "область", value)
            value = re.sub(r"г\.([^\s]+)", r"город \1", value)
            value = re.sub(r"г\.\s+([^\s]+)", r"город \1", value)
            value = re.sub(r"[Рр]есп\.", "Республика", value)
            value = re.sub(r"с\.", "село", value)
            value = re.sub(r"им\.", "имени", value)
            value = re.sub(r"\"", " ", value)
            value = re.sub(r"«", " ", value)
            value = re.sub(r"»", " ", value)
            value = " ".join(w.strip() for w in value.split())
            
            print(f"{emc}. {value}")
            


if __name__ == "__main__":
    main()
